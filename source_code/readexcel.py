import openpyxl

#workbook object
wb=openpyxl.Workbook()

sheet=wb.active

c1=sheet.cell(row=1,column=1)
c1.value="Sunil"

c2=sheet.cell(row=1,column=2)
c2.value="Anil"

wb.save("D:\\test.xlsx")






